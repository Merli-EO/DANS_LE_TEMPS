from openpyxl import load_workbook
from pathlib import Path
from .models import Story
import datetime

def Data_Creation() :
    # Path to the excel file
    static_dir_api = Path(__file__).resolve().parent
    excel_file_path = static_dir_api / 'static/Dans_Le_Temps_notebook.xlsx'

    # Load the excel file
    workbook = load_workbook(filename=excel_file_path)

    # Select the active sheet
    sheet = workbook.active
    Story.objects.all().delete()
    row_number = 0
    for row in sheet.iter_rows():

        row_number += 1
        # Ignorer la première ligne
        if row_number == 1:
            continue
        # Create the model object


        story = Story(
            title = row[0].value,
            publication_date = datetime.datetime.now(), #row[1].value,
            url =   row[2].value,
            moral = row[3].value,
            description = row[4].value,
            thumbnail = 'http://ec2-16-171-173-230.eu-north-1.compute.amazonaws.com:8000/static/Dans_LE_Temps_images/' + row[5].value
            )
        print(story.title)
        story.save()
